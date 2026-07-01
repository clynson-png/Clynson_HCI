import os
import sqlite3
import openpyxl
import tempfile
import shutil
import subprocess
import time
from datetime import datetime

# Configurações do ambiente (baseadas no hci_admin_control.py)
ADB = os.path.expandvars(r"%LOCALAPPDATA%\Android\Sdk\platform-tools\adb.exe")
PACKAGE = "com.example.sportsperformance"
EXCEL_PATH = r"C:/Users/clyns/AndroidStudioProjects/SportsPerformance/app/src/main/assets/999_HCI_V18_VBA_ESTAVEL - Copia.xlsm"

def run_adb(args):
    cmd = [ADB, *args]
    return subprocess.run(cmd, capture_output=True, check=True)

def copy_app_database():
    print("Baixando banco de dados do app...")
    tmp = tempfile.mkdtemp(prefix="hci-import-db-")
    db_path = os.path.join(tmp, "sports_performance_db")
    for suffix in ["", "-wal", "-shm"]:
        remote = f"databases/sports_performance_db{suffix}"
        local = os.path.join(tmp, f"sports_performance_db{suffix}")
        try:
            result = run_adb(["exec-out", "run-as", PACKAGE, "cat", remote])
            with open(local, "wb") as fh:
                fh.write(result.stdout)
        except subprocess.CalledProcessError:
            if suffix == "": raise
    return tmp, db_path

def push_database(tmp):
    print("Enviando banco de dados atualizado para o app...")
    run_adb(["shell", "am", "force-stop", PACKAGE])
    for suffix in ["", "-wal", "-shm"]:
        local = os.path.join(tmp, f"sports_performance_db{suffix}")
        if os.path.exists(local):
            remote_tmp = f"/data/local/tmp/sports_performance_db{suffix}"
            run_adb(["push", local, remote_tmp])
            run_adb(["shell", "run-as", PACKAGE, "cp", remote_tmp, f"databases/sports_performance_db{suffix}"])
            run_adb(["shell", "rm", "-f", remote_tmp])
    print("Banco de dados atualizado com sucesso!")

def import_data():
    tmp_dir, db_path = copy_app_database()

    print(f"Abrindo Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "INPUT" not in wb.sheetnames:
        print("ERRO: Aba 'INPUT' não encontrada no Excel.")
        return

    sheet = wb["INPUT"]
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    # Limpa tabela atual para evitar duplicidade se desejar um refresh total,
    # ou apenas usa INSERT OR REPLACE. Vou usar INSERT OR REPLACE.

    rows_processed = 0
    header = [str(cell.value) for cell in sheet[1]]

    # Mapeamento de colunas (0-indexed)
    # DATA_COLETA(0), PROVA(1), ATLETA(2), EVENTO(3), SESSAO(4), ID_BLOCO(5), STATUS_EVENTO(6), SERIE(8), T1(9)...T10(18)

    for row in sheet.iter_rows(min_row=3): # Começa na linha 3 (pulando cabeçalho e linha vazia)
        vals = [c.value for c in row]
        if not vals[2]: continue # Se não tem atleta, pula

        data_coleta = vals[0]
        if isinstance(data_coleta, datetime):
            timestamp = int(data_coleta.timestamp() * 1000)
        else:
            timestamp = int(time.time() * 1000)

        prova = str(vals[1] or "PISTOL").upper()
        atleta = str(vals[2])
        evento = str(vals[3] or "EV1")
        sessao = str(vals[4] or "TREINO")
        id_bloco = str(vals[5] or "BLOCO1")
        status_evento = str(vals[6] or "PARCIAL")
        serie = str(vals[8] or "SR1")

        # Tiros T1 a T10 (colunas 9 a 18)
        tiros_list = []
        for i in range(9, 19):
            val = vals[i]
            if val is not None:
                try: tiros_list.append(float(val))
                except: pass

        tiros_str = ",".join(map(str, tiros_list))

        # hciSerieOrder
        try:
            order = int("".join(filter(str.isdigit, serie)) or "1")
        except:
            order = 1

        # chaveSerie: {athlete}_{prova}_{event}_{session}_{id_bloco}_{serie}
        chave = f"{atleta}_{prova}_{evento}_{sessao}_{id_bloco}_{serie}"

        cur.execute("""
            INSERT OR REPLACE INTO shot_series
            (chaveSerie, dataColeta, prova, atleta, evento, sessao, idBloco, statusEvento, serie, tiros, hciSerieOrder, hciEventRowValid)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (chave, timestamp, prova, atleta, evento, sessao, id_bloco, status_evento, serie, tiros_str, order, 1))

        rows_processed += 1

    conn.commit()
    conn.close()

    push_database(tmp_dir)
    shutil.rmtree(tmp_dir)
    print(f"Importação concluída: {rows_processed} séries inseridas.")

if __name__ == "__main__":
    try:
        import_data()
    except Exception as e:
        print(f"ERRO durante a importação: {e}")
